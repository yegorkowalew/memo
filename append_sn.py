# -*- coding: utf-8 -*-
"""Append data from files to DB"""

import os
import time
from datetime import datetime
from threading import Thread
import logging
logger = logging.getLogger(__name__)

import pandas as pd
from openpyxl import load_workbook

import platform
import django
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "memo.settings")
django.setup()
from service_note.models import ServiceNote, OperativeNote, ReadyOrder, InDocument, Waybill, SheduleFile, SheduleFileError, Shedule

IN_DOCUMENT_FOLDER = 'C:\\work\\memo\\График документации'
IN_DOCUMENT_FILE = 'График документации.xlsx'

def insertToDB(df_records, m_obj):
    """
    Вставляем данные в БД
    """
    model_instances = [m_obj(**record) for record in df_records]
    try:
        m_obj.objects.bulk_create(model_instances)
        logger.debug('paste')
    except BaseException as identifier:
        logger.error("insertToDB error with class: %s, %s" % (m_obj.__class__.__name__, identifier))

"""
Служебные записки
ServiceNote
"""
SN_FILE = 'C:\\work\\memo\\Служебные записки v2.xlsx'

def serviceNoteReadFile(sn_file):
    try:
        df = pd.read_excel(
            sn_file, 
            sheet_name="СЗ",
            parse_dates = [
                'Отгрузка "от"',
                'Отгрузка "до"',
                'Дата СЗ',
                'Дата СЗ Факт',
                'Комплектовочные План Дата',
                'Отгрузочные План Дата',
                'Конструкторская документация План Дата',
                'Материалы План Дата',
                'Черный металл План Дата',
                'Оцинкованный металл План Дата',
                'Чугун План Дата',
            ],
            dtype={
                'ID':int,
                'Продукция':str,
                'Контрагент':str,
                '№ Заказа':str,
                '№ СЗ с изменениями':str,
            }
        )

    except Exception as ind:
        logger.error("serviceNoteReadFile error with file: %s - %s" % (sn_file, ind))
    else:
        df = df.rename(columns={
            'ID':'in_id',
            'Порядковый номер':'in_sn_no',
            'Отгрузка "от"':'shipment_from',
            'Отгрузка "до"':'shipment_before',
            'Продукция':'product_name',
            'Контрагент':'counterparty',
            '№ Заказа':'order_no',
            'Кол-во':'amount',
            '№ СЗ':'sn_no',
            '№ СЗ с изменениями':'sn_no_amended',
            'Дата СЗ':'sn_date',
            'Дата СЗ Факт':'sn_date_fact',
            'Комплектовочные План Дней':'pickup_plan_days',
            'Комплектовочные План Дата':'pickup_plan_date',
            'Отгрузочные План Дней':'shipping_plan_days',
            'Отгрузочные План Дата':'shipping_plan_date',
            'Конструкторская документация План Дней':'design_plan_days',
            'Конструкторская документация План Дата':'design_plan_date',
            'Материалы План Дней':'material_plan_days',
            'Материалы План Дата':'material_plan_date',
            'Черный металл План Дней':'black_metal_plan_days',
            'Черный металл План Дата':'black_metal_plan_date',
            'Оцинкованный металл План Дней':'galvanized_metal_plan_days',
            'Оцинкованный металл План Дата':'galvanized_metal_plan_date',
            'Чугун План Дней':'cast_iron_plan_days',
            'Чугун План Дата':'cast_iron_plan_date',
            }
        )
        df = df.astype(object)
        df = df.where(df.notnull(), None)
        df_records = df.to_dict('records')
        return df_records

def serviceNoteRebuild(sn_file, m_obj):
    t = time.time()
    m_obj.objects.all().delete()
    logger.info("Clear DBTable: {} - {:.3}".format(m_obj.__name__, (time.time()-t)))
    t = time.time()
    df = serviceNoteReadFile(sn_file)
    logger.info("Convert to Table: {} - {:.3}".format(m_obj.__name__, (time.time()-t)))
    t = time.time()
    insertToDB(df, m_obj)
    logger.info("Insert to DBTable: {} - {:.3}".format(m_obj.__name__, (time.time()-t)))
    
"""
Готовые заказы
ReadyOrder
"""
RO_FILE = 'C:\\work\\memo\\Готовые заказы.xlsx'

def readyOrderReadFile(ro_file):
    try:
        df = pd.read_excel(
            ro_file, 
            sheet_name="Готовые заказы",
            usecols = [
                'ID',
                'Готово',
                'Готово Дата'
            ],
            parse_dates = [
                'Готово Дата',
            ],
            dtype={
                'ID':int,
            },
            converters={
                'Готово':bool,
            }
        )

    except Exception as ind:
        logger.error("readyOrderReadFile error with file: %s - %s" % (ro_file, ind))
    else:
        df = df.rename(columns={
            'ID':'in_id',
            'Готово':'ready',
            'Готово Дата':'ready_date',
            }
        )
        df.fillna({'':False})
        df["ready_date"] = pd.to_datetime(df["ready_date"])
        df = df.astype({
            'in_id': 'int64',
            'ready':'bool',
            'ready_date':'object',
            })
        df = df.where(df.notnull(), None)
        df_records = df.to_dict('records')
        return df_records

def readyOrderRebuild(sn_file, m_obj):
    t = time.time()
    m_obj.objects.all().delete()
    logger.info("Clear DBTable: {} - {:.3}".format(m_obj.__name__, (time.time()-t)))
    t = time.time()
    df = readyOrderReadFile(sn_file)
    logger.info("Convert to Table: {} - {:.3}".format(m_obj.__name__, (time.time()-t)))
    t = time.time()
    insertToDB(df, m_obj)
    logger.info("Insert to DBTable: {} - {:.3}".format(m_obj.__name__, (time.time()-t)))


"""
График документации
"""

def inDocumentReadFile(path, file_name):
    
    try:
        df = pd.read_excel(
            os.path.join(path, file_name),
            sheet_name="Документация",
            usecols = [
                'ID',
                'Комплектовочные Выдача',
                'Комплектовочные Дата 1',
                'Комплектовочные Дата 2',
                'Комплектовочные Дата 3',
                'Отгрузочные Выдача',
                'Отгрузочные Дата 1',
                'Отгрузочные Дата 2',
                'Отгрузочные Дата 3',
                'Конструкторские Выдача',
                'Конструкторские Дата 1',
                'Конструкторские Дата 2',
                'Конструкторские Дата 3',
                'Изменения чертежей Выдача',
                'Изменения чертежей Дата 1',
                'Изменения чертежей Дата 2',
                'Изменения чертежей Дата 3'
            ],

            parse_dates = [
                'Комплектовочные Дата 1',
                'Комплектовочные Дата 2',
                'Комплектовочные Дата 3',
                'Отгрузочные Дата 1',
                'Отгрузочные Дата 2',
                'Отгрузочные Дата 3',
                'Конструкторские Дата 1',
                'Конструкторские Дата 2',
                'Конструкторские Дата 3',
                'Изменения чертежей Дата 1',
                'Изменения чертежей Дата 2',
                'Изменения чертежей Дата 3'
            ],

            dtype={
                'ID':int,
            },
            converters={
                'Готово':bool,
                'Комплектовочные Выдача':bool,
                'Отгрузочные Выдача':bool,
                'Конструкторские Выдача':bool,
                'Изменения чертежей Выдача':bool,
            }
        )

    except Exception as ind:
        logger.error("inDocumentReadFile error with file: %s - %s" % (file_name, ind))
    else:
        df = df.rename(columns={
            'ID':'in_id',
            'Диспетчер':'dispatcher',
            'Комплектовочные Выдача':'pickup_issue',
            'Комплектовочные Дата 1':'pickup_date_1',
            'Комплектовочные Дата 2':'pickup_date_2',
            'Комплектовочные Дата 3':'pickup_date_3',
            'Отгрузочные Выдача':'shipping_issue',
            'Отгрузочные Дата 1':'shipping_date_1',
            'Отгрузочные Дата 2':'shipping_date_2',
            'Отгрузочные Дата 3':'shipping_date_3',
            'Конструкторские Выдача':'design_issue',
            'Конструкторские Дата 1':'design_date_1',
            'Конструкторские Дата 2':'design_date_2',
            'Конструкторские Дата 3':'design_date_3',
            'Изменения чертежей Выдача':'drawings_issue',
            'Изменения чертежей Дата 1':'drawings_date_1',
            'Изменения чертежей Дата 2':'drawings_date_2',
            'Изменения чертежей Дата 3':'drawings_date_3'
            }
        )
        df['dispatcher'] = os.path.split(path)[-1]
        df['pickup_date_1'] = pd.to_datetime(df['pickup_date_1'])
        df['pickup_date_2'] = pd.to_datetime(df['pickup_date_2'])
        df['pickup_date_3'] = pd.to_datetime(df['pickup_date_3'])
        df['shipping_date_1'] = pd.to_datetime(df['shipping_date_1'])
        df['shipping_date_2'] = pd.to_datetime(df['shipping_date_2'])
        df['shipping_date_3'] = pd.to_datetime(df['shipping_date_3'])
        df['design_date_1'] = pd.to_datetime(df['design_date_1'])
        df['design_date_2'] = pd.to_datetime(df['design_date_2'])
        df['design_date_3'] = pd.to_datetime(df['design_date_3'])
        df['drawings_date_1'] = pd.to_datetime(df['drawings_date_1'])
        df['drawings_date_2'] = pd.to_datetime(df['drawings_date_2'])
        df['drawings_date_3'] = pd.to_datetime(df['drawings_date_3'])

        df = df.astype({
            'pickup_date_1':'object',
            'pickup_date_2':'object',
            'pickup_date_3':'object',
            'shipping_date_1':'object',
            'shipping_date_2':'object',
            'shipping_date_3':'object',
            'design_date_1':'object',
            'design_date_2':'object',
            'design_date_3':'object',
            'drawings_date_1':'object',
            'drawings_date_2':'object',
            'drawings_date_3':'object',
            })
        df = df.where(df.notnull(), None)
        df = df.to_dict('records')
        return df 

def inDocumentFindFile(in_folder, need_file):
    tree = os.walk(in_folder)
    folder_list = []
    for i in tree:
        for address, dirs, files in [i]:
            for fl in files:
                if fl == need_file:
                    folder_list.append(address)
    return folder_list

def inDocumentRebuild(in_folder, need_file, m_obj):
    t = time.time()
    m_obj.objects.all().delete()
    logger.info("Clear DBTable: {} - {:.3}".format(m_obj.__name__, (time.time()-t)))

    for folder in inDocumentFindFile(in_folder, need_file):
        t = time.time()
        df = inDocumentReadFile(folder, need_file)
        logger.info("Convert to Table: {} - {:.3}".format(m_obj.__name__, (time.time()-t)))
        t = time.time()
        insertToDB(df, m_obj)
        logger.info("Insert to DBTable: {} - {:.3}".format(m_obj.__name__, (time.time()-t)))

"""
Накладные
"""

N_FOLDER = 'C:\\work\\memo\\График накладных\\Выгрузки'

def waybillReadFile(files):
    # Документ
    # Номер
    # Дата
    # Проведен
    # ЗаказНомер
    # Склад
    # ЦехПолучатель
    # СкладПолучатель
    # Номенклатура
    # ХарактеристикаНоменклатуры
    # Количество
    # ЕдиницаИзмерения
    # Заказ
    # Ответственный
    # Коэффициент
    # Ссылка

    df_list = [pd.read_csv(
        file, encoding = 'Windows-1251', 
        sep='\t',
        dtype={
                'Документ':str,
                'Номер':str,
                'Дата':str,
                'Проведен':str,
                'ЗаказНомер':str,
                'Склад':str,
                'ЦехПолучатель':str,
                'СкладПолучатель':str,
                'Номенклатура':str,
                'ХарактеристикаНоменклатуры':str,
                'Количество':str,
                'ЕдиницаИзмерения':str,
                'Заказ':str,
                'Ответственный':str,
                'Коэффициент':str,
                'Ссылка':str,
            },
            parse_dates = [
                'Дата',
            ],
    
    ) for file in files]

    big_df = pd.concat(df_list)

    big_df = big_df.rename(columns={
        'Документ':'document',
        'Номер':'number',
        'Дата':'date',
        'Проведен':'held',
        'ЗаказНомер':'order_no',
        'Склад':'stock',
        'ЦехПолучатель':'shop_recipient',
        'СкладПолучатель':'warehouse_recipient',
        'Номенклатура':'nomenclature',
        'ХарактеристикаНоменклатуры':'item_feature',
        'Количество':'amount',
        'ЕдиницаИзмерения':'measure',
        'Заказ':'order_no_text',
        'Ответственный':'responsible',
        'Коэффициент':'coefficient',
        'Ссылка':'link',
        })

    import numpy as np

    import locale
    from locale import atof
    locale.setlocale(locale.LC_NUMERIC, '')

    big_df['coefficient'] = big_df['coefficient'].apply(locale.atof).astype(float)
    big_df['amount'] = big_df['amount'].apply(locale.atof).astype(float)
    big_df["held"] = np.where(big_df["held"] == "Да", True, False).astype(bool)
    big_df = big_df.where(big_df.notnull(), None)
    df_records = big_df.to_dict('records')
    return df_records

def waybillFindFile(in_folder):
    extentions = ['txt']
    needfiles = []
    for root, subfolders, nfiles in os.walk(in_folder):
        for filee in nfiles:
            if filee.split('.')[-1] in extentions:
                needfiles.append(os.path.join(root, filee))
    return needfiles

def waybillRebuild(in_folder, m_obj):
    t = time.time()
    m_obj.objects.all().delete()
    logger.info("Clear DBTable: {} - {:.3}".format(m_obj.__name__, (time.time()-t)))
    t = time.time()
    files = waybillFindFile(in_folder)
    df = waybillReadFile(files)
    logger.info("Convert to Table: {} - {:.3}".format(m_obj.__name__, (time.time()-t)))
    t = time.time()
    insertToDB(df, m_obj)
    logger.info("Insert to DBTable: {} - {:.3}".format(m_obj.__name__, (time.time()-t)))

def allRebuild():
    try:
        # t = time.time()
        readyOrderRebuild(RO_FILE, ReadyOrder)
        serviceNoteRebuild(SN_FILE, ServiceNote)
        inDocumentRebuild(IN_DOCUMENT_FOLDER, IN_DOCUMENT_FILE, InDocument)
        waybillRebuild(N_FOLDER, Waybill)
        shedule_files_rebuild(SHEDULE_FOLDER)

        # print(time.time()-t)
        # 4.3612494468688965
        # 4.120489597320557

        # t = time.time()
        # t1 = Thread(target=readyOrderRebuild, args=(RO_FILE, ReadyOrder))
        # t2 = Thread(target=serviceNoteRebuild, args=(SN_FILE, ServiceNote))
        # t3 = Thread(target=inDocumentRebuild, args=(IN_DOCUMENT_FOLDER, IN_DOCUMENT_FILE, InDocument))
        # t4 = Thread(target=waybillRebuild, args=(N_FOLDER, Waybill))
        # t5 = Thread(target=shedule_files_rebuild, args=(SHEDULE_FOLDER))
        
        # t1.start()
        # t2.start()
        # t3.start()
        # t4.start()
        # t5.start()

        # t1.join()
        # t2.join()
        # t3.join()
        # t4.join()
        # t5.join()
        
        # print(time.time()-t)
        # 3.922224283218384
        # 4.084655284881592

        # t = time.time()
        # t1 = Process(target=readyOrderRebuild, args=(RO_FILE, ReadyOrder))
        # t2 = Process(target=serviceNoteRebuild, args=(SN_FILE, ServiceNote))
        # t3 = Process(target=inDocumentRebuild, args=(IN_DOCUMENT_FOLDER, IN_DOCUMENT_FILE, InDocument))
        # t1.start()
        # t2.start()
        # t3.start()
        # t1.join()
        # t2.join()
        # t3.join()
        # print(time.time()-t)
        # 4.757272243499756
        # 5.010553359985352

        # t = time.time()
        # t1 = multiprocessing.Process(target=readyOrderRebuild, args=(RO_FILE, ReadyOrder))
        # t2 = multiprocessing.Process(target=serviceNoteRebuild, args=(SN_FILE, ServiceNote))
        # t3 = multiprocessing.Process(target=inDocumentRebuild, args=(IN_DOCUMENT_FOLDER, IN_DOCUMENT_FILE, InDocument))
        # t1.start()
        # t2.start()
        # t3.start()
        # t1.join()
        # t2.join()
        # t3.join()
        # print(time.time()-t)
        # 8.132465124130249
        # 4.950515508651733

        # t = time.time()
        # p = Pool(3)
        # t1 = p.apply_async(readyOrderRebuild,(RO_FILE, ReadyOrder))
        # t2 = p.apply_async(serviceNoteRebuild,(SN_FILE, ServiceNote))
        # t3 = p.apply_async(inDocumentRebuild,(IN_DOCUMENT_FOLDER, IN_DOCUMENT_FILE, InDocument))
        # p.close()
        # p.join()
        # print(time.time()-t)
        # 4.860277891159058
        # 5.0001044273376465
        return True
    except:
        return False

"""
Список файлов графиков ПДО
"""
SHEDULE_FOLDER = 'C:\\work\\memo\\Графики ПДО'

def shedule_find_file(in_folder):
    extentions = ['xlsx']
    needfiles = []
    for root, _, nfiles in os.walk(in_folder):
        for filee in nfiles:
            if filee.split('.')[-1] in extentions:
                needfiles.append(os.path.join(root, filee))
    return needfiles

def creation_date(path_to_file):
    if platform.system() == 'Windows':
        creation_date = datetime.utcfromtimestamp(os.path.getctime(path_to_file))
        modification_date = datetime.utcfromtimestamp(os.path.getmtime(path_to_file))
        return creation_date, modification_date
    else:
        stat = os.stat(path_to_file)
        try:
            return stat.st_birthtime
        except AttributeError:
            creation_date = datetime.utcfromtimestamp(stat.st_mtime)
            return creation_date, None

# def get_dispatcher_from_path(file_path):
#     zz = file_path.split('\\')[-3].split(' - ')[-1]
#     print(os.path.split(file_path))
#     return zz

def get_dispatcher_from_path(file_path):
    return file_path.partition('Диспетчер')[2].partition(' - ')[2].partition('\\')[0]

def shedule_files_read(files):
    """
    Парсим имена графиков
    """
    error_files = []
    correct_files = []
    for xls_file in files:
        try:
            wb = load_workbook(filename=xls_file, read_only=True)
            ws = wb['График']
            cell_value = ws.cell(row=3, column=3).value
            if len(str(cell_value)) != 7:
                # print(xls_file)
                error_files.append({
                    # 'dispatcher':xls_file.split('\\')[-3].split(' - ')[-1],
                    'dispatcher':get_dispatcher_from_path(xls_file),
                    'error':'Order number: %s in file does not match valid number' % (cell_value),
                    'file_path':xls_file
                })
            else:
                c_m_dates = creation_date(xls_file)
                correct_files.append({
                    # 'dispatcher':xls_file.split('\\')[-2].split(' - ')[-1],
                    'dispatcher':get_dispatcher_from_path(xls_file),
                    'file_path':xls_file,
                    'order_no':cell_value,
                    'date_creation': c_m_dates[0],
                    'date_modification': c_m_dates[1],
                    })
        except BaseException as identifier:
            # print(xls_file.split('\\')[-2].split(' - ')[-2])
            # print(xls_file.split('\\')[-3].split(' - ')[-1])
            error_files.append({
                # 'dispatcher':xls_file.split('\\')[-3].split(' - ')[-1],
                'dispatcher':get_dispatcher_from_path(xls_file),
                'file_path':xls_file,
                'error':identifier,
            })
    return correct_files, error_files

def shedule_files_rebuild(in_folder):
    t = time.time()
    SheduleFile.objects.all().delete()
    logger.info("Clear DBTable: {} - {:.3}".format(SheduleFile.__name__, (time.time()-t)))

    t = time.time()
    SheduleFileError.objects.all().delete()
    logger.info("Clear DBTable: {} - {:.3}".format(SheduleFileError.__name__, (time.time()-t)))

    t = time.time()
    xls_files = shedule_find_file(in_folder)
    correct_files, error_files = shedule_files_read(xls_files)
    logger.info("Convert to Table: {} - {:.3}".format('Shedule Files', (time.time()-t)))

    t = time.time()
    insertToDB(correct_files, SheduleFile)
    logger.info("Insert to Table: {} - {:.3}".format(SheduleFile.__name__, (time.time()-t)))

    t = time.time()
    insertToDB(error_files, SheduleFileError)
    logger.info("Insert to Table: {} - {:.3}".format(SheduleFileError.__name__, (time.time()-t)))

def shedule_files_table_to_df(xls_file_path):
    data = pd.read_excel(
        xls_file_path,
        sheet_name="График",
        header=None
    )
    order_no = data.loc[2,2]
    headers = data.iloc[5].replace(to_replace=r'\n', value=' ', regex=True)
    new_table = pd.DataFrame(data.values[6:], columns=headers)
    new_table.insert(0, "Заказ №", order_no)
    # big_df = pd.concat(df_list)
    new_table.rename(
        columns={
            'Заказ №':'order_no',
            'Узел':'knot',
            'Наименование':'name',
            'Операция':'operation',
            'Кол-во Узел':'amount_knot',
            'Кол-во Изделие':'amount_product',
            'Кол-во Заказ':'amount_order',
            'Кол-во Массив':'amount_array',
            'Материал':'material',
            'Размер':'size',
            'Размер заготовки':'size_workpiece',
            'Росцеховка':'shop_path',
        }, 
        inplace=True
    )
    # print('Parse %s' % (xls_file_path))
    # print('>>>')
    return new_table

def shedule_files_table_parse(files):
    for xls_file in files:
        try:
            # print('Parse %s' % (xls_file))
            df = shedule_files_table_to_df(xls_file['file_path'])
            big_df = df.where(df.notnull(), None)
            df_records = big_df.to_dict('records')
            insertToDB(df_records, Shedule)
        except BaseException as identifier:
            print(f'Trabl: {identifier}')
            time.sleep(10)
        
    return 'df_records'

if __name__ == "__main__":
    t = time.time()
    # shedule_files_rebuild(SHEDULE_FOLDER)
    # get_dispatcher_from_path('C:\work\memo\Графики ПДО\Диспетчер-2 - Мудренко Наталія Володимирівна\22 -ТОВ Авис Зернотрейд\Новый шаблон.xlsx')

    allRebuild()
    # xls_files = shedule_find_file(SHEDULE_FOLDER)
    # correct_files, error_files = shedule_files_read(xls_files)

    Shedule.objects.all().delete()
    xls_files = shedule_find_file(SHEDULE_FOLDER)
    correct_files, error_files = shedule_files_read(xls_files)
    records = shedule_files_table_parse(correct_files)

    print("Общее время:")
    print(time.time()-t)
